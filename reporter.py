import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               numbers as oxl_numbers)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart, Series
from openpyxl.drawing.image import Image as XLImage

OUT = Path(__file__).parent.parent / "outputs"
CHARTS = OUT / "charts"

# ── Color palette ──────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", start_color="1F4E79")   # dark navy
HDR2_FILL  = PatternFill("solid", start_color="2E75B6")   # medium blue
PASS_FILL  = PatternFill("solid", start_color="C6EFCE")   # light green
FAIL_FILL  = PatternFill("solid", start_color="FFC7CE")   # light red
ALT_FILL   = PatternFill("solid", start_color="EBF3FB")   # very light blue
WARN_FILL  = PatternFill("solid", start_color="FFEB9C")   # yellow
SECTION_FILL = PatternFill("solid", start_color="D6E4F0") # section header

HDR_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HDR2_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
BOLD_FONT  = Font(name="Arial", bold=True, size=9)
NORMAL_FONT= Font(name="Arial", size=9)
TITLE_FONT = Font(name="Arial", bold=True, size=12, color="1F4E79")
SEC_FONT   = Font(name="Arial", bold=True, size=9, color="1F4E79")

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=False)
RIGHT  = Alignment(horizontal="right",  vertical="center")

def _style_cell(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font:      cell.font = font
    if fill:      cell.fill = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border = border
    if number_format: cell.number_format = number_format

def _header_row(ws, row, cols_data, hdr_fill=None, hdr_font=None):
    hfill = hdr_fill or HDR_FILL
    hfont = hdr_font or HDR_FONT
    for col_idx, (label, width) in enumerate(cols_data, start=1):
        c = ws.cell(row=row, column=col_idx, value=label)
        _style_cell(c, font=hfont, fill=hfill, alignment=CENTER, border=THIN_BORDER)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def _data_row(ws, row, values, alt=False, fills=None):
    fill = ALT_FILL if alt else PatternFill("solid", start_color="FFFFFF")
    for col_idx, val in enumerate(values, start=1):
        c = ws.cell(row=row, column=col_idx, value=val)
        cell_fill = fills[col_idx-1] if (fills and fills[col_idx-1]) else fill
        _style_cell(c, font=NORMAL_FONT, fill=cell_fill,
                    alignment=CENTER, border=THIN_BORDER)


# ──────────────────────────────────────────────────────────────
# SHEET 1 – CBR_Raw
# ──────────────────────────────────────────────────────────────
def _write_cbr_raw(wb, df_cbr):
    ws = wb.create_sheet("CBR_Raw")
    ws.row_dimensions[1].height = 20
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "CBR Raw Data — DCP Measurement per Station"
    _style_cell(t, font=TITLE_FONT, alignment=LEFT)

    cols = [("Road", 18), ("Station", 14), ("CBR Loaded (%)", 16),
            ("CBR Standard (%)", 17), ("Status", 12)]
    _header_row(ws, 2, cols)

    for i, row in enumerate(df_cbr.itertuples(), start=3):
        ok = row.cbr_ok
        stat_fill = PASS_FILL if ok else FAIL_FILL
        fills = [None, None, None, None, stat_fill]
        _data_row(ws, i, [row.road, row.station_label, row.cbr_loaded,
                          row.cbr_standard, "✓ OK" if ok else "✗ FAIL"],
                  alt=(i % 2 == 0), fills=fills)
    ws.freeze_panes = "A3"
    return ws


# ──────────────────────────────────────────────────────────────
# SHEET 2 – CBR_Summary
# ──────────────────────────────────────────────────────────────
def _write_cbr_summary(wb, df_cbr_sum):
    ws = wb.create_sheet("CBR_Summary")
    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = "CBR Summary by Road — vs Standard (39% for HD 785)"
    _style_cell(t, font=TITLE_FONT, alignment=LEFT)

    cols = [("Road", 16), ("Stations", 12), ("OK", 8), ("Fail", 8),
            ("Fail %", 10), ("Min CBR", 10), ("Max CBR", 10),
            ("Mean CBR", 11), ("Standard", 11)]
    _header_row(ws, 2, cols)

    for i, row in enumerate(df_cbr_sum.itertuples(), start=3):
        fail_fill = FAIL_FILL if row.pct_fail > 0 else PASS_FILL
        fills = [None]*4 + [fail_fill] + [None]*4
        _data_row(ws, i,
                  [row.road, row.n_stations, row.n_ok, row.n_fail,
                   f"{row.pct_fail:.1f}%", row.cbr_min_actual,
                   row.cbr_max_actual, row.cbr_mean_actual, row.cbr_standard],
                  fills=fills)
    return ws


# ──────────────────────────────────────────────────────────────
# SHEET 3 – Grade_Raw
# ──────────────────────────────────────────────────────────────
def _write_grade_raw(wb, df_grade):
    ws = wb.create_sheet("Grade_Raw")
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = "Grade Raw Data — per Station (Standard ≤ 8%)"
    _style_cell(t, font=TITLE_FONT, alignment=LEFT)

    cols = [("Road", 18), ("Station", 14), ("Grade (%)", 14),
            ("Standard (%)", 14), ("Status", 12)]
    _header_row(ws, 2, cols)

    for i, row in enumerate(df_grade.itertuples(), start=3):
        ok = row.grade_ok
        stat_fill = PASS_FILL if ok else FAIL_FILL
        fills = [None]*4 + [stat_fill]
        _data_row(ws, i, [row.road, row.station_label,
                          round(row.grade_pct, 2), row.grade_standard,
                          "✓ OK" if ok else "✗ OVER"],
                  alt=(i % 2 == 0), fills=fills)
    ws.freeze_panes = "A3"
    return ws


# ──────────────────────────────────────────────────────────────
# SHEET 4 – RDS_Eval
# ──────────────────────────────────────────────────────────────
def _write_rds(wb, df_rds_detail, df_rds_summary):
    ws = wb.create_sheet("RDS_Eval")
    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = "Qualitative Rolling Resistance Assessment — Thompson (2011)"
    _style_cell(t, font=TITLE_FONT, alignment=LEFT)

    # Detail
    ws["A2"] = "Detail per Defect"
    _style_cell(ws["A2"], font=SEC_FONT, fill=SECTION_FILL, alignment=LEFT)
    ws.merge_cells("A2:F2")

    cols = [("Road", 16), ("Defect", 18), ("Degree (1-5)", 14),
            ("Extent (1-5)", 14), ("Defect Score", 14), ("Contribution %", 15)]
    _header_row(ws, 3, cols, hdr_fill=HDR2_FILL, hdr_font=HDR2_FONT)

    defect_colors = {
        "potholes": "FCE4D6", "corrugations": "FDEBD0",
        "rutting": "FEF9E7", "loose_material": "E8F8F5", "stoniness": "EBF5FB"
    }
    for i, row in enumerate(df_rds_detail.itertuples(), start=4):
        dc = defect_colors.get(row.defect, "FFFFFF")
        df = PatternFill("solid", start_color=dc)
        fills = [None, df, None, None, None, None]
        _data_row(ws, i, [row.road, row.defect, row.degree, row.extent,
                          row.defect_score, f"{row.contribution_pct:.1f}%"],
                  alt=(i % 2 == 0), fills=fills)

    # Summary
    sum_start = 4 + len(df_rds_detail) + 2
    ws.cell(sum_start, 1).value = "Summary per Road"
    ws.merge_cells(f"A{sum_start}:F{sum_start}")
    _style_cell(ws.cell(sum_start, 1), font=SEC_FONT, fill=SECTION_FILL, alignment=LEFT)

    cols2 = [("Road", 16), ("Total RDS", 14), ("RR (%)", 12),
             ("RR Target (%)", 15), ("RR OK?", 12), ("Dominant Defect", 18)]
    _header_row(ws, sum_start + 1, cols2, hdr_fill=HDR2_FILL, hdr_font=HDR2_FONT)

    for i, row in enumerate(df_rds_summary.itertuples(), start=sum_start + 2):
        ok = row.rr_ok
        stat_fill = PASS_FILL if ok else FAIL_FILL
        fills = [None]*4 + [stat_fill] + [None]
        _data_row(ws, i, [row.road, row.total_rds, row.rr_pct, row.rr_target,
                          "✓ OK" if ok else "✗ FAIL", row.dominant_defect],
                  fills=fills)
    return ws


# ──────────────────────────────────────────────────────────────
# SHEET 5 – Giroud_Han
# ──────────────────────────────────────────────────────────────
def _write_giroud_han(wb, df_gh):
    ws = wb.create_sheet("Giroud_Han")
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "Pavement Design — Giroud-Han Method (Tebal Base Course per Station)"
    _style_cell(t, font=TITLE_FONT, alignment=LEFT)

    cols = [("Road", 16), ("Station", 14), ("CBR Subgrade (%)", 18),
            ("r Contact (m)", 14), ("Ph=0 (kN)", 14),
            ("RE", 10), ("Fe", 10), ("Base Course h (m)", 18)]
    _header_row(ws, 2, cols)

    for i, row in enumerate(df_gh.itertuples(), start=3):
        h = row.base_course_h_m
        h_fill = WARN_FILL if h > 0.55 else (PASS_FILL if h <= 0.50 else None)
        fills = [None]*7 + [h_fill]
        _data_row(ws, i, [row.road, row.station_label, row.cbr_loaded,
                          row.r_m, row.Ph0_kN, row.RE, row.Fe, h],
                  alt=(i % 2 == 0), fills=fills)

    # annotation
    note_row = 3 + len(df_gh) + 1
    ws.cell(note_row, 1).value = (
        "Color key: Green = h ≤ 0.50m  |  Yellow = h > 0.55m  |  "
        "Standard: CBR min 39%,  Nc=3.14 (no geotextile),  rutting allow 75mm"
    )
    _style_cell(ws.cell(note_row, 1), font=Font(name="Arial", italic=True, size=8, color="595959"))
    ws.merge_cells(f"A{note_row}:H{note_row}")
    ws.freeze_panes = "A3"
    return ws


# ──────────────────────────────────────────────────────────────
# SHEET 6 – Rimpull_Speed
# ──────────────────────────────────────────────────────────────
def _write_rimpull(wb, df_rimpull, df_speed_summary):
    ws = wb.create_sheet("Rimpull_Speed")
    ws.merge_cells("A1:M1")
    t = ws["A1"]
    t.value = "Rimpull & Speed Analysis — Before vs After Repair (Loaded Direction)"
    _style_cell(t, font=TITLE_FONT, alignment=LEFT)

    cols = [
        ("Road", 16), ("Station", 14),
        ("Grade Before (%)", 17), ("Grade After (%)", 16),
        ("RR Before (%)", 14), ("RR After (%)", 13),
        ("TR Before (lb)", 16), ("TR After (lb)", 15),
        ("Gear Before", 13), ("Gear After", 12),
        ("Speed Before (km/h)", 20), ("Speed After (km/h)", 19),
        ("Δ Speed (km/h)", 15),
    ]
    _header_row(ws, 2, cols)

    for i, row in enumerate(df_rimpull.itertuples(), start=3):
        delta = row.speed_delta_kph
        delta_fill = PASS_FILL if delta > 0 else (FAIL_FILL if delta < 0 else None)
        fills = [None]*12 + [delta_fill]
        _data_row(ws, i, [
            row.road, row.station_label,
            row.grade_before_pct, row.grade_after_pct,
            row.RR_before_pct, row.RR_after_pct,
            row.TR_before_lb, row.TR_after_lb,
            row.gear_before, row.gear_after,
            row.speed_before_kph, row.speed_after_kph,
            delta,
        ], alt=(i % 2 == 0), fills=fills)

    # Summary block
    sum_start = 3 + len(df_rimpull) + 2
    ws.cell(sum_start, 1).value = "Speed Summary"
    ws.merge_cells(f"A{sum_start}:E{sum_start}")
    _style_cell(ws.cell(sum_start, 1), font=SEC_FONT, fill=SECTION_FILL)

    sum_cols = [("Road", 16), ("Avg Speed Actual (km/h)", 22),
                ("Avg Speed Improved (km/h)", 23), ("Δ Avg (km/h)", 14)]
    _header_row(ws, sum_start+1, sum_cols[:4], hdr_fill=HDR2_FILL, hdr_font=HDR2_FONT)

    for i, row in enumerate(df_speed_summary.itertuples(), start=sum_start+2):
        _data_row(ws, i, [row.road, row.speed_actual_avg,
                          row.speed_improved_avg, row.speed_delta_avg])

    ws.freeze_panes = "A3"
    return ws


# ──────────────────────────────────────────────────────────────
# SHEET 7 – Cut_Fill
# ──────────────────────────────────────────────────────────────
def _write_cut_fill(wb, df_cf, df_cf_summary):
    ws = wb.create_sheet("Cut_Fill")
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "Cut & Fill Volume Estimation — Grade Before vs After Repair (per 50m Station)"
    _style_cell(t, font=TITLE_FONT, alignment=LEFT)

    cols = [("Road", 16), ("Station", 14), ("Grade Before (%)", 17),
            ("Grade After (%)", 16), ("Δ Grade (%)", 13),
            ("Δh (m)", 10), ("Volume (m³)", 14), ("Work Type", 12)]
    _header_row(ws, 2, cols)

    for i, row in enumerate(df_cf.itertuples(), start=3):
        wt = row.work_type
        wt_fill = FAIL_FILL if wt == "cut" else (PASS_FILL if wt == "fill" else None)
        fills = [None]*7 + [wt_fill]
        _data_row(ws, i, [row.road, row.station_label, row.grade_before_pct,
                          row.grade_after_pct, row.delta_grade_pct,
                          row.dh_m, row.volume_m3, row.work_type.upper()],
                  alt=(i % 2 == 0), fills=fills)

    # Summary
    sum_start = 3 + len(df_cf) + 2
    ws.cell(sum_start, 1).value = "Volume Summary per Road"
    ws.merge_cells(f"A{sum_start}:F{sum_start}")
    _style_cell(ws.cell(sum_start, 1), font=SEC_FONT, fill=SECTION_FILL)

    sum_cols = [("Road", 16), ("Total Cut (m³)", 16), ("Total Fill (m³)", 16),
                ("Net (m³)", 13), ("Cut Stations", 14), ("Fill Stations", 14)]
    _header_row(ws, sum_start+1, sum_cols, hdr_fill=HDR2_FILL, hdr_font=HDR2_FONT)
    for i, row in enumerate(df_cf_summary.itertuples(), start=sum_start+2):
        _data_row(ws, i, [row.road, row.total_cut_m3, row.total_fill_m3,
                          row.net_m3, row.n_cut_stations, row.n_fill_stations])

    ws.freeze_panes = "A3"
    return ws


# ──────────────────────────────────────────────────────────────
# SHEET 8 – Method_Notes
# ──────────────────────────────────────────────────────────────
def _write_method_notes(wb):
    ws = wb.create_sheet("Method_Notes")
    notes = [
        ("Project 2 — Road Damage Analyzer", "PT PPA Adaro Indonesia, Pit Wara, Tabalong, Kalimantan Selatan"),
        ("Reference Paper", "Figgo Febriawan et al., Jurnal Teknologi Pertambangan Vol.10 No.2, Jan 2025"),
        ("", ""),
        ("PHASE 1 — CBR Evaluation", ""),
        ("Method", "DCP (Dynamic Cone Penetrometer) per 50m station, loaded lane"),
        ("Standard", "CBR min = 39% (required for Komatsu HD 785, beban 170 ton)"),
        ("Basis", "Distribusi beban roda depan = 63,713 lb → daya dukung 8.6 kg/cm²"),
        ("", ""),
        ("PHASE 1 — Grade Evaluation", ""),
        ("Method", "Survey kemiringan per 50m station menggunakan software Civil 3D"),
        ("Standard", "Grade max = 8% (TS-AI-PRO-06-003)"),
        ("", ""),
        ("PHASE 2 — RDS Scoring", ""),
        ("Method", "Qualitative Rolling Resistance Assessment — Thompson (2011)"),
        ("Defects scored", "Potholes, Corrugations, Rutting, Loose Material, Stoniness"),
        ("Formula", "Defect Score = Degree (1-5) × Extent (1-5); Total RDS → lookup RR%"),
        ("RR Target", "≤ 2% (company standard PT PPA)"),
        ("", ""),
        ("PHASE 3 — Pavement Design", ""),
        ("Method", "Giroud-Han (Suwandhi, 2004) — iterasi nonlinear via scipy.optimize.fsolve"),
        ("Key parameters", "Nc=3.14 (no geotextile), rutting izin=75mm, traffic=1500 kend/hari"),
        ("Base material", "EWT Floor 100 dari Pit Wara, CBR ≥ 39%"),
        ("", ""),
        ("PHASE 4 — Rimpull & Speed", ""),
        ("Method", "Chart Performance Komatsu HD 785-7 (Power Mode)"),
        ("Baseline", "TR dan gear aktual dari Tabel 14-15 jurnal Figo 2024"),
        ("Improved speed", "Setelah perbaikan grade + RR=2%, TR dihitung ulang → gear → speed"),
        ("", ""),
        ("PHASE 4 — Cut & Fill", ""),
        ("Method", "Estimasi volumetrik per station: Δh = Δgrade/100 × 50m"),
        ("Formula", "Volume = |Δh| × lebar jalan (25m) × panjang station (50m)"),
        ("Basis grade", "Tabel 19-20 jurnal Figo 2024 (before & after repair)"),
        ("Note", "Volume total cross-check vs Civil 3D: MonteBawah fill~14562m³, cut~1677m³; Spanyol fill~6924m³, cut~6809m³"),
    ]

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 72

    ws.merge_cells("A1:B1")
    ws["A1"] = "Method Notes & Assumptions — Project 2 Road Damage Analyzer"
    _style_cell(ws["A1"], font=Font(name="Arial", bold=True, size=13, color="1F4E79"),
                alignment=LEFT)
    ws.row_dimensions[1].height = 24

    for i, (key, val) in enumerate(notes, start=2):
        ck = ws.cell(i, 1, key)
        cv = ws.cell(i, 2, val)
        if key and not val:
            ck.font = SEC_FONT
            ck.fill = SECTION_FILL
            ws.merge_cells(f"A{i}:B{i}")
            _style_cell(ck, font=SEC_FONT, fill=SECTION_FILL, alignment=LEFT)
        elif key and val:
            _style_cell(ck, font=BOLD_FONT, alignment=LEFT, border=THIN_BORDER)
            _style_cell(cv, font=NORMAL_FONT, alignment=LEFT, border=THIN_BORDER)
        ws.row_dimensions[i].height = 15

    return ws


# ──────────────────────────────────────────────────────────────
# CHART helpers (matplotlib → png → embedded)
# ──────────────────────────────────────────────────────────────
def plot_cbr_profile(df_cbr, cbr_min):
    CHARTS.mkdir(parents=True, exist_ok=True)
    roads = df_cbr["road"].unique()
    fig, axes = plt.subplots(1, len(roads), figsize=(14, 5), sharey=False)
    if len(roads) == 1: axes = [axes]
    colors = {"MonteBawah": "#2980B9", "Spanyol": "#E74C3C"}
    for ax, road in zip(axes, roads):
        sub = df_cbr[df_cbr["road"] == road]
        bar_colors = [("#27AE60" if ok else "#E74C3C") for ok in sub["cbr_ok"]]
        ax.bar(range(len(sub)), sub["cbr_loaded"], color=bar_colors, edgecolor="white", width=0.7)
        ax.axhline(cbr_min, color="navy", linestyle="--", lw=1.5, label=f"Standard {cbr_min}%")
        ax.set_xticks(range(len(sub)))
        ax.set_xticklabels(sub["station_label"], rotation=45, ha="right", fontsize=7)
        ax.set_title(road, fontweight="bold", fontsize=11)
        ax.set_ylabel("CBR (%)")
        ax.legend(fontsize=8)
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    fig.suptitle("CBR Profile per Station — Actual vs Standard (39%)", fontweight="bold", fontsize=12)
    plt.tight_layout()
    path = CHARTS / "cbr_profile.png"
    plt.savefig(path, dpi=150, bbox_inches="tight"); plt.close()
    print(f"[reporter] Chart saved: {path}")

def plot_grade_profile(df_grade, grade_max):
    CHARTS.mkdir(parents=True, exist_ok=True)
    roads = df_grade["road"].unique()
    fig, axes = plt.subplots(len(roads), 1, figsize=(14, 8), sharex=False)
    if len(roads) == 1: axes = [axes]
    for ax, road in zip(axes, roads):
        sub = df_grade[df_grade["road"] == road]
        bar_colors = [("#27AE60" if ok else "#E74C3C") for ok in sub["grade_ok"]]
        ax.bar(range(len(sub)), sub["grade_pct"].abs(), color=bar_colors, edgecolor="white", width=0.7)
        ax.axhline(grade_max, color="navy", linestyle="--", lw=1.5, label=f"Max {grade_max}%")
        ax.set_xticks(range(len(sub)))
        ax.set_xticklabels(sub["station_label"], rotation=45, ha="right", fontsize=7)
        ax.set_title(road, fontweight="bold", fontsize=10)
        ax.set_ylabel("Grade (%)")
        ax.legend(fontsize=8)
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    fig.suptitle("Grade Profile per Station — Actual vs Standard (8%)", fontweight="bold", fontsize=12)
    plt.tight_layout()
    path = CHARTS / "grade_profile.png"
    plt.savefig(path, dpi=150, bbox_inches="tight"); plt.close()
    print(f"[reporter] Chart saved: {path}")

def plot_speed_comparison(df_rimpull, df_speed_summary):
    CHARTS.mkdir(parents=True, exist_ok=True)
    roads = df_rimpull["road"].unique()
    fig, axes = plt.subplots(len(roads), 1, figsize=(14, 9))
    if len(roads) == 1: axes = [axes]
    for ax, road in zip(axes, roads):
        sub = df_rimpull[df_rimpull["road"] == road].reset_index(drop=True)
        x = np.arange(len(sub))
        w = 0.35
        b1 = ax.bar(x - w/2, sub["speed_before_kph"], w, label="Aktual (Sebelum)",
                    color="#E74C3C", alpha=0.85, edgecolor="white")
        b2 = ax.bar(x + w/2, sub["speed_after_kph"], w, label="Setelah Perbaikan",
                    color="#27AE60", alpha=0.85, edgecolor="white")

        avg_b = df_speed_summary.loc[df_speed_summary["road"]==road, "speed_actual_avg"].values[0]
        avg_a = df_speed_summary.loc[df_speed_summary["road"]==road, "speed_improved_avg"].values[0]
        ax.axhline(avg_b, color="#C0392B", linestyle="--", lw=1.2, alpha=0.7,
                   label=f"Avg Aktual {avg_b:.1f} km/h")
        ax.axhline(avg_a, color="#1E8449", linestyle="--", lw=1.2, alpha=0.7,
                   label=f"Avg Perbaikan {avg_a:.1f} km/h")

        ax.set_xticks(x); ax.set_xticklabels(sub["station_label"], rotation=45, ha="right", fontsize=7)
        ax.set_ylabel("Speed (km/h)"); ax.set_title(road, fontweight="bold", fontsize=11)
        ax.legend(fontsize=8, ncol=2)
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
        ax.set_ylim(0, 50)

    fig.suptitle("Speed Comparison — Before vs After Road Repair (HD 785 Loaded)", fontweight="bold", fontsize=12)
    plt.tight_layout()
    path = CHARTS / "speed_comparison.png"
    plt.savefig(path, dpi=150, bbox_inches="tight"); plt.close()
    print(f"[reporter] Chart saved: {path}")

def plot_cut_fill_chart(df_cf):
    CHARTS.mkdir(parents=True, exist_ok=True)
    roads = df_cf["road"].unique()
    fig, axes = plt.subplots(1, len(roads), figsize=(14, 5))
    if len(roads) == 1: axes = [axes]
    for ax, road in zip(axes, roads):
        sub = df_cf[df_cf["road"] == road].reset_index(drop=True)
        colors = ["#E74C3C" if t == "cut" else ("#27AE60" if t == "fill" else "#95A5A6")
                  for t in sub["work_type"]]
        ax.bar(range(len(sub)), sub["volume_m3"], color=colors, edgecolor="white", width=0.7)
        ax.set_xticks(range(len(sub)))
        ax.set_xticklabels(sub["station_label"], rotation=45, ha="right", fontsize=7)
        ax.set_ylabel("Volume (m³)"); ax.set_title(road, fontweight="bold", fontsize=11)
        cut_p = mpatches.Patch(color="#E74C3C", label="Cut")
        fill_p = mpatches.Patch(color="#27AE60", label="Fill")
        ax.legend(handles=[cut_p, fill_p], fontsize=9)
        ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    fig.suptitle("Cut & Fill Volume per Station — Grade Repair", fontweight="bold", fontsize=12)
    plt.tight_layout()
    path = CHARTS / "cut_fill.png"
    plt.savefig(path, dpi=150, bbox_inches="tight"); plt.close()
    print(f"[reporter] Chart saved: {path}")


# ──────────────────────────────────────────────────────────────
# MAIN write_report
# ──────────────────────────────────────────────────────────────
def write_report(df_cbr, df_grade, cbr_sum, grade_sum,
                 df_rds_detail=None, df_rds_summary=None,
                 df_gh=None,
                 df_rimpull=None, df_speed_summary=None,
                 df_cf=None, df_cf_summary=None):

    OUT.mkdir(parents=True, exist_ok=True)
    filepath = OUT / "report.xlsx"

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    _write_cbr_raw(wb, df_cbr)
    _write_cbr_summary(wb, cbr_sum)
    _write_grade_raw(wb, df_grade)

    if df_rds_detail is not None:
        _write_rds(wb, df_rds_detail, df_rds_summary)

    if df_gh is not None:
        _write_giroud_han(wb, df_gh)

    if df_rimpull is not None:
        _write_rimpull(wb, df_rimpull, df_speed_summary)

    if df_cf is not None:
        _write_cut_fill(wb, df_cf, df_cf_summary)

    _write_method_notes(wb)

    wb.save(filepath)
    print(f"[reporter] Report saved: {filepath}")
